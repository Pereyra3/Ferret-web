from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook

from core.models import Store
from warehouse.models import Product, StockLevel, StockMovement, StockTransfer
from sales.models import Sale
from warehouse.models import Purchase
from warehouse.services.stock import (
    accept_transfer,
    apply_adjustment,
    apply_purchase,
    apply_sale,
    apply_transfer,
    reject_transfer,
    user_can_accept_transfer,
)
from warehouse.services.stock_import import (
    StockImportError,
    apply_stock_import,
    parse_stock_rows,
)


def _xlsx_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.mark.django_db
class TestStockServiceEdges:
    def test_apply_sale_skips_draft_and_idempotent(
        self, store_ops_setup, create_confirmed_sale
    ):
        data = store_ops_setup
        draft = Sale.objects.create(
            store=data["store"], user=data["user"], status=Sale.Status.DRAFT
        )
        apply_sale(draft, data["user"])
        sale = create_confirmed_sale()
        apply_sale(sale, data["user"])
        data["stock"].refresh_from_db()
        qty = data["stock"].quantity
        apply_sale(sale, data["user"])
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == qty

    def test_apply_purchase_idempotent(self, store_ops_setup, create_purchase):
        data = store_ops_setup
        purchase = create_purchase()
        apply_purchase(purchase, data["user"])
        purchase.refresh_from_db()
        assert purchase.stock_applied
        apply_purchase(purchase, data["user"])

    def test_apply_purchase_already_applied_flag(self, store_ops_setup, create_purchase):
        purchase = create_purchase()
        purchase.stock_applied = True
        purchase.save(update_fields=["stock_applied"])
        apply_purchase(purchase, store_ops_setup["user"])

    def test_apply_transfer_zero_qty(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Z", code="z")
        transfer = StockTransfer.objects.create(
            from_store=data["store"], to_store=dest, user=data["user"]
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("0"))
        with pytest.raises(ValueError, match="inválida"):
            apply_transfer(transfer, data["user"])

    def test_apply_transfer_already_applied_early(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Y", code="y")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
            applied=True,
            status=StockTransfer.Status.ACCEPTED,
        )
        apply_transfer(transfer, data["user"])


@pytest.mark.django_db
class TestStockTransfer:
    def test_apply_transfer_moves_stock(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Norte", code="norte")
        product = data["product"]
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=product, quantity=Decimal("10"))
        apply_transfer(transfer, data["user"])
        transfer.refresh_from_db()
        assert transfer.applied is True
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == Decimal("90")
        dest_level = StockLevel.objects.get(store=dest, product=product)
        assert dest_level.quantity == Decimal("10")
        assert StockMovement.objects.filter(
            reason=StockMovement.Reason.TRANSFER_OUT
        ).exists()

    def test_apply_transfer_same_store_raises(self, store_ops_setup):
        data = store_ops_setup
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=data["store"],
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("1"))
        with pytest.raises(ValueError, match="distintas"):
            apply_transfer(transfer, data["user"])

    def test_apply_transfer_no_origin_stock(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Vacía", code="vacia")
        other = Product.objects.create(sku="OTRO-1", name="Otro")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=other, quantity=Decimal("1"))
        with pytest.raises(ValueError, match="Sin existencia"):
            apply_transfer(transfer, data["user"])

    def test_apply_transfer_idempotent(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Idem", code="idem")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("1"))
        apply_transfer(transfer, data["user"])
        data["stock"].refresh_from_db()
        qty_after = data["stock"].quantity
        apply_transfer(transfer, data["user"])
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == qty_after

    def test_apply_transfer_insufficient_stock(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Sur", code="sur")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("500"))
        with pytest.raises(ValueError, match="insuficiente"):
            apply_transfer(transfer, data["user"])

    def test_transfer_create_view(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Este", code="este")
        dest.assigned_users.add(data["user"])
        response = authenticated_client.post(
            reverse("stock_transfer_create"),
            {
                "from_store": data["store"].pk,
                "to_store": dest.pk,
                "notes": "Traslado",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-0-product": data["product"].pk,
                "lines-0-quantity": "5",
            },
        )
        assert response.status_code == 302
        transfer = StockTransfer.objects.get()
        assert transfer.status == StockTransfer.Status.PENDING
        assert transfer.applied is False

    def test_transfer_approval_flow(self, client, store_ops_setup, create_user):
        from django.contrib.auth.models import Group

        from core.roles import GROUP_ENCARGADO

        data = store_ops_setup
        dest = Store.objects.create(name="Este", code="este2")
        dest.assigned_users.add(data["user"])
        receiver = create_user(username="receiver", password="pass")
        receiver.groups.set([Group.objects.get(name=GROUP_ENCARGADO)])
        receiver.save()
        dest.assigned_users.add(receiver)
        client.login(username="testuser", password="password")
        client.post(
            reverse("stock_transfer_create"),
            {
                "from_store": data["store"].pk,
                "to_store": dest.pk,
                "notes": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-0-product": data["product"].pk,
                "lines-0-quantity": "5",
            },
        )
        transfer = StockTransfer.objects.get()
        client.logout()
        client.login(username="receiver", password="pass")
        response = client.post(reverse("stock_transfer_accept", args=[transfer.pk]))
        assert response.status_code == 302
        transfer.refresh_from_db()
        assert transfer.applied is True
        assert transfer.status == StockTransfer.Status.ACCEPTED


@pytest.mark.django_db
class TestStockImport:
    def test_parse_and_apply_set_mode(self, store_ops_setup):
        data = store_ops_setup
        raw = _xlsx_bytes(
            [
                ["sku", "cantidad"],
                [data["product"].sku, 42],
            ]
        )
        rows, _, _ = parse_stock_rows(BytesIO(raw))
        applied = apply_stock_import(
            data["store"], data["user"], rows, mode="set"
        )
        assert applied == 1
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == Decimal("42")

    def test_parse_invalid_binary(self):
        with pytest.raises(StockImportError, match="No se pudo leer"):
            parse_stock_rows(BytesIO(b"not-an-xlsx"))

    def test_parse_header_only_no_products(self):
        raw = _xlsx_bytes([["sku", "cantidad"]])
        with pytest.raises(StockImportError, match="No hay filas"):
            parse_stock_rows(BytesIO(raw))

    def test_parse_missing_qty_on_row(self):
        raw = _xlsx_bytes([["sku", "cantidad"], ["SOLO-SKU"]])
        with pytest.raises(StockImportError, match="Fila 2"):
            parse_stock_rows(BytesIO(raw))

    def test_parse_header_with_none_sku_label(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([None, "cantidad"])
        sheet.append(["X-1", 2])
        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        with pytest.raises(StockImportError, match="SKU"):
            parse_stock_rows(buf)

    def test_parse_empty_workbook_sheet(self):
        workbook = Workbook()
        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        with pytest.raises(StockImportError, match="vacío"):
            parse_stock_rows(buf)

    def test_parse_negative_qty(self):
        raw = _xlsx_bytes([["sku", "cantidad"], ["A", -1]])
        with pytest.raises(StockImportError):
            parse_stock_rows(BytesIO(raw))

    def test_parse_skips_fully_empty_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["sku", "cantidad"])
        sheet.append([None, None])
        sheet.append(["ROW-OK", 7])
        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        rows, _, _ = parse_stock_rows(buf)
        assert len(rows) == 1

    def test_parse_skips_blank_sku_row(self):
        raw = _xlsx_bytes(
            [
                ["sku", "cantidad"],
                ["", 5],
                ["ONLY", 3],
            ]
        )
        rows, _, _ = parse_stock_rows(BytesIO(raw))
        assert len(rows) == 1

    def test_apply_invalid_mode(self, store_ops_setup):
        data = store_ops_setup
        with pytest.raises(ValueError):
            apply_stock_import(data["store"], data["user"], [], mode="bad")

    def test_apply_skips_zero_delta(self, store_ops_setup):
        data = store_ops_setup
        raw = _xlsx_bytes(
            [["sku", "cantidad"], [data["product"].sku, data["stock"].quantity]]
        )
        rows, _, _ = parse_stock_rows(BytesIO(raw))
        assert apply_stock_import(data["store"], data["user"], rows, mode="set") == 0

    def test_parse_invalid_qty(self):
        raw = _xlsx_bytes([["sku", "cantidad"], ["A", "abc"]])
        with pytest.raises(StockImportError):
            parse_stock_rows(BytesIO(raw))

    def test_parse_missing_sku_column(self):
        raw = _xlsx_bytes([["nombre", "qty"], ["x", 1]])
        with pytest.raises(StockImportError, match="SKU"):
            parse_stock_rows(BytesIO(raw))

    def test_import_view(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        content = _xlsx_bytes(
            [
                ["codigo", "existencia"],
                [data["product"].sku, 50],
            ]
        )
        upload = SimpleUploadedFile(
            "inv.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = authenticated_client.post(
            reverse("stock_import"),
            {"file": upload, "mode": "set"},
        )
        assert response.status_code == 302
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == Decimal("50")

    def test_import_sample_download(self, authenticated_client):
        response = authenticated_client.get(reverse("stock_import_sample"))
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_apply_add_mode(self, store_ops_setup):
        data = store_ops_setup
        raw = _xlsx_bytes(
            [["sku", "cantidad"], [data["product"].sku, 10]]
        )
        rows, _, _ = parse_stock_rows(BytesIO(raw))
        apply_stock_import(data["store"], data["user"], rows, mode="add")
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == Decimal("110")

    def test_unknown_sku_raises(self, store_ops_setup):
        data = store_ops_setup
        raw = _xlsx_bytes([["sku", "cantidad"], ["NO-EXISTE", 1]])
        rows, _, _ = parse_stock_rows(BytesIO(raw))
        with pytest.raises(StockImportError, match="no existe"):
            apply_stock_import(data["store"], data["user"], rows, mode="set")

    def test_import_view_invalid_file(self, authenticated_client):
        upload = SimpleUploadedFile("bad.xlsx", b"not excel", content_type="application/octet-stream")
        response = authenticated_client.post(
            reverse("stock_import"),
            {"file": upload, "mode": "set"},
        )
        assert response.status_code == 200

    def test_transfer_same_store_rejected(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        response = authenticated_client.post(
            reverse("stock_transfer_create"),
            {
                "from_store": data["store"].pk,
                "to_store": data["store"].pk,
                "notes": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-0-product": data["product"].pk,
                "lines-0-quantity": "1",
            },
        )
        assert response.status_code == 200

    def test_transfer_insufficient_on_accept(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="Lejos", code="lejos")
        dest.assigned_users.add(data["user"])
        authenticated_client.post(
            reverse("stock_transfer_create"),
            {
                "from_store": data["store"].pk,
                "to_store": dest.pk,
                "notes": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-0-product": data["product"].pk,
                "lines-0-quantity": "99999",
            },
        )
        transfer = StockTransfer.objects.get()
        response = authenticated_client.post(
            reverse("stock_transfer_accept", args=[transfer.pk])
        )
        assert response.status_code == 302
        transfer.refresh_from_db()
        assert transfer.status == StockTransfer.Status.PENDING
        data["stock"].refresh_from_db()
        assert data["stock"].quantity == Decimal("100")

    def test_models_str(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="X", code="x")
        transfer = StockTransfer.objects.create(
            from_store=data["store"], to_store=dest, user=data["user"]
        )
        line = transfer.lines.create(product=data["product"], quantity=Decimal("1"))
        assert "Transfer" in str(transfer)
        assert data["product"].sku in str(line)


@pytest.mark.django_db
class TestTransferApprovalService:
    def test_reject_and_accept_guards(self, store_ops_setup, create_user):
        data = store_ops_setup
        dest = Store.objects.create(name="R", code="r")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("1"))
        dest.assigned_users.add(data["user"])
        reject_transfer(transfer, data["user"])
        transfer.refresh_from_db()
        assert transfer.status == StockTransfer.Status.REJECTED
        reject_transfer(transfer, data["user"])
        with pytest.raises(ValueError, match="pendientes"):
            accept_transfer(transfer, data["user"])
        with pytest.raises(ValueError, match="rechazada"):
            apply_transfer(transfer, data["user"])

    def test_reject_without_permission(self, store_ops_setup, create_user):
        data = store_ops_setup
        dest = Store.objects.create(name="P", code="p")
        other = create_user(username="other", password="pass")
        transfer = StockTransfer.objects.create(
            from_store=data["store"], to_store=dest, user=data["user"]
        )
        with pytest.raises(ValueError, match="permiso"):
            reject_transfer(transfer, other)

    def test_user_can_accept_transfer(self, store_ops_setup, create_user):
        data = store_ops_setup
        dest = Store.objects.create(name="C", code="c")
        transfer = StockTransfer.objects.create(
            from_store=data["store"], to_store=dest, user=data["user"]
        )
        assert user_can_accept_transfer(data["user"], transfer) is False
        dest.assigned_users.add(data["user"])
        assert user_can_accept_transfer(data["user"], transfer) is True

    def test_accept_already_accepted(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="A", code="a")
        dest.assigned_users.add(data["user"])
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
            status=StockTransfer.Status.ACCEPTED,
            applied=True,
        )
        accept_transfer(transfer, data["user"])

    def test_superuser_can_accept_any_store(self, store_ops_setup, create_user):
        data = store_ops_setup
        dest = Store.objects.create(name="S", code="s")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        admin = create_user(username="admin", password="pass")
        admin.is_superuser = True
        admin.save()
        assert user_can_accept_transfer(admin, transfer) is True

    def test_reject_non_pending_raises(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="NP", code="np")
        dest.assigned_users.add(data["user"])
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
            status=StockTransfer.Status.ACCEPTED,
        )
        with pytest.raises(ValueError, match="pendientes"):
            reject_transfer(transfer, data["user"])

    def test_reject_pending_changed_in_db(self, store_ops_setup):
        from unittest.mock import MagicMock, patch

        data = store_ops_setup
        dest = Store.objects.create(name="DB", code="db")
        dest.assigned_users.add(data["user"])
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        locked = StockTransfer.objects.get(pk=transfer.pk)
        locked.status = StockTransfer.Status.ACCEPTED
        mock_qs = MagicMock()
        mock_qs.get.return_value = locked
        with patch.object(
            StockTransfer.objects, "select_for_update", return_value=mock_qs
        ):
            with pytest.raises(ValueError, match="pendientes"):
                reject_transfer(transfer, data["user"])

    def test_accept_without_permission_raises(self, store_ops_setup, create_user):
        data = store_ops_setup
        dest = Store.objects.create(name="AP", code="ap")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("1"))
        other = create_user(username="noperms", password="pass")
        with pytest.raises(ValueError, match="permiso"):
            accept_transfer(transfer, other)

    def test_apply_non_pending_raises(self, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="AP2", code="ap2")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
            status=StockTransfer.Status.ACCEPTED,
            applied=False,
        )
        with pytest.raises(ValueError, match="pendiente"):
            apply_transfer(transfer, data["user"])


@pytest.mark.django_db
class TestTransferApprovalViews:
    def test_list_and_reject(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="L", code="l")
        dest.assigned_users.add(data["user"])
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        transfer.lines.create(product=data["product"], quantity=Decimal("2"))
        body = authenticated_client.get(reverse("stock_transfer_list")).content.decode()
        assert "Por recibir" in body
        response = authenticated_client.post(
            reverse("stock_transfer_reject", args=[transfer.pk])
        )
        assert response.status_code == 302
        transfer.refresh_from_db()
        assert transfer.status == StockTransfer.Status.REJECTED

    def test_accept_denied_wrong_store(self, client, store_ops_setup, create_user):
        from django.contrib.auth.models import Group

        from core.roles import GROUP_ENCARGADO

        data = store_ops_setup
        dest = Store.objects.create(name="D", code="d")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        outsider = create_user(username="outsider", password="pass")
        outsider.groups.set([Group.objects.get(name=GROUP_ENCARGADO)])
        outsider.save()
        client.login(username="outsider", password="pass")
        response = client.post(reverse("stock_transfer_accept", args=[transfer.pk]))
        assert response.status_code == 302
        transfer.refresh_from_db()
        assert transfer.status == StockTransfer.Status.PENDING

    def test_reject_denied_wrong_store(self, client, store_ops_setup, create_user):
        from django.contrib.auth.models import Group

        from core.roles import GROUP_ENCARGADO

        data = store_ops_setup
        dest = Store.objects.create(name="RD", code="rd")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        outsider = create_user(username="outsider_rej", password="pass")
        outsider.groups.set([Group.objects.get(name=GROUP_ENCARGADO)])
        outsider.save()
        client.login(username="outsider_rej", password="pass")
        response = client.post(reverse("stock_transfer_reject", args=[transfer.pk]))
        assert response.status_code == 302
        transfer.refresh_from_db()
        assert transfer.status == StockTransfer.Status.PENDING

    def test_reject_view_value_error(self, authenticated_client, store_ops_setup):
        data = store_ops_setup
        dest = Store.objects.create(name="RV", code="rv")
        dest.assigned_users.add(data["user"])
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
            status=StockTransfer.Status.ACCEPTED,
        )
        response = authenticated_client.post(
            reverse("stock_transfer_reject", args=[transfer.pk])
        )
        assert response.status_code == 302
        assert transfer.status == StockTransfer.Status.ACCEPTED

    def test_cajero_cannot_accept(self, client, django_user_model, store_ops_setup):
        from django.contrib.auth.models import Group
        from django.core.management import call_command

        from core.roles import GROUP_CAJERO

        call_command("setup_roles")
        data = store_ops_setup
        dest = Store.objects.create(name="Caj", code="caj")
        transfer = StockTransfer.objects.create(
            from_store=data["store"],
            to_store=dest,
            user=data["user"],
        )
        cajero = django_user_model.objects.create_user(
            username="cajero_xfer", password="pass"
        )
        cajero.groups.set([Group.objects.get(name=GROUP_CAJERO)])
        cajero.save()
        client.login(username="cajero_xfer", password="pass")
        assert (
            client.post(reverse("stock_transfer_accept", args=[transfer.pk])).status_code
            == 403
        )
