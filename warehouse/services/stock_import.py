"""Load stock levels from Excel (.xlsx)."""

from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from warehouse.models import Product
from warehouse.services.stock import apply_adjustment

SKU_HEADERS = frozenset({"sku", "codigo", "código", "clave", "barcode"})
QTY_HEADERS = frozenset({"cantidad", "quantity", "qty", "existencia", "stock"})


class StockImportError(Exception):
    """Row-level or file-level import errors with a user-facing message."""


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_decimal(value) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise StockImportError("Cantidad vacía.")
    try:
        qty = Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise StockImportError(f"Cantidad inválida: {value!r}") from exc
    if qty < 0:
        raise StockImportError("La cantidad no puede ser negativa.")
    return qty


def parse_stock_rows(uploaded_file):
    """
    Read first sheet; return list of {sku, quantity} and detected column labels.
    First row must contain SKU and quantity column headers.
    """
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise StockImportError(
            "No se pudo leer el archivo. Use Excel .xlsx válido."
        ) from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise StockImportError("El archivo está vacío.")
    header = [_normalize_header(cell) for cell in rows[0]]
    sku_col = None
    qty_col = None
    for idx, label in enumerate(header):
        if label in SKU_HEADERS:
            sku_col = idx
        if label in QTY_HEADERS:
            qty_col = idx
    if sku_col is None or qty_col is None:
        raise StockImportError(
            "La primera fila debe incluir columnas SKU (o código) y cantidad "
            "(o existencia)."
        )
    parsed = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        sku_raw = row[sku_col] if sku_col < len(row) else None
        sku = str(sku_raw).strip() if sku_raw is not None else ""
        if not sku:
            continue
        qty_cell = row[qty_col] if qty_col < len(row) else None
        try:
            quantity = _parse_decimal(qty_cell)
        except StockImportError as exc:
            raise StockImportError(f"Fila {row_num}: {exc}") from exc
        parsed.append({"sku": sku, "quantity": quantity, "row": row_num})
    if not parsed:
        raise StockImportError("No hay filas de producto para importar.")
    return parsed, header[sku_col], header[qty_col]


@transaction.atomic
def apply_stock_import(store, user, rows, *, mode: str):
    """
    mode 'set': cantidad = existencia final.
    mode 'add': cantidad = delta a sumar.
    """
    if mode not in ("set", "add"):
        raise ValueError("mode must be 'set' or 'add'")
    applied = 0
    for row in rows:
        product = Product.objects.filter(sku__iexact=row["sku"]).first()
        if product is None:
            raise StockImportError(
                f"Fila {row['row']}: producto con SKU «{row['sku']}» no existe."
            )
        target = row["quantity"]
        if mode == "set":
            current = product.stock_quantity(store)
            delta = Decimal(target) - Decimal(current)
        else:
            delta = Decimal(target)
        if delta == 0:
            continue
        apply_adjustment(
            store,
            product,
            delta,
            user,
            reference=f"import:{mode}",
        )
        applied += 1
    return applied
